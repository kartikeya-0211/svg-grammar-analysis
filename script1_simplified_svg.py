"""
SCRIPT 1: RE-FETCH, PROCESS & VISUALIZE (SMART LOCAL LOAD)
-------------------------------------------------
1. Reads URLs from 'railroad_diagrams.xlsx' (Column B).
2. Checks if local SVG exists. If not, RE-FETCHES from web.
3. Updates Column C (Raw SVG Code) and Column D (Raw SVG Image).
4. Checks if redlines SVG exists. If not, processes it.
5. Updates Column E (Connected SVG Code) and Column F (Connected SVG Image).
6. Saves files to 'rawSVGs' and 'redlinesSVGs' folders.
"""

import os
import re
import time
import math
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options

# --- CONFIGURATION ---
INPUT_FILE = 'railroad_diagrams.xlsx'  # Must contain URLs in Col B
DRIVER_FILENAME = "msedgedriver.exe"
DRIVER_PATH = os.path.join(os.getcwd(), DRIVER_FILENAME)
RAW_DIR = 'rawSVGs'
REDLINES_DIR = 'redlinesSVGs'

# ==========================================
# PART 1: SELENIUM SETUP & FETCHING
# ==========================================
def setup_driver():
    if not os.path.exists(DRIVER_PATH):
        raise Exception(f"❌ msedgedriver.exe not found at: {DRIVER_PATH}")

    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument("--log-level=3") 
    options.add_argument('--disable-gpu')
    options.add_argument('--force-device-scale-factor=1')
    options.add_argument('--window-size=2500,2500') 
    
    # Block images to speed up the re-fetch process
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    
    service = Service(DRIVER_PATH)
    service.creation_flags = 0x08000000 
    return webdriver.Edge(service=service, options=options)

def fetch_svg_source_live(driver, url):
    """Goes to the URL and grabs the fresh SVG code."""
    try:
        driver.get(url)
        wait = WebDriverWait(driver, 5)
        svg_elem = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "svg.syntaxdiagram")))
        return svg_elem.get_attribute("outerHTML")
    except:
        return None

def take_screenshot(driver, svg_string, output_path):
    """Renders SVG string in a temp file and takes a screenshot."""
    if not svg_string: return False
    try:
        html = f'''<html><body style="margin: 0; padding: 10px; background: white;">{svg_string}</body></html>'''
        temp_html = os.path.abspath("temp_render.html")
        with open(temp_html, "w", encoding="utf-8") as f:
            f.write(html)
            
        driver.get(f"file:///{temp_html}")
        time.sleep(0.15) # Brief wait for render
        
        svg = driver.find_element(By.TAG_NAME, "svg")
        svg.screenshot(output_path)
        return True
    except: return False

# ==========================================
# PART 2: FLATTENING & GEOMETRY
# ==========================================
def calculate_bounds(root):
    min_x, min_y = 10000.0, 10000.0
    max_x, max_y = 0.0, 0.0
    found = False
    for elem in root.iter():
        xs, ys = [], []
        tag = elem.tag.split('}')[-1]
        
        if tag == 'line':
            xs = [float(elem.get('x1',0)), float(elem.get('x2',0))]
            ys = [float(elem.get('y1',0)), float(elem.get('y2',0))]
        elif tag in ['rect','text']:
            try:
                x, y = float(elem.get('x',0)), float(elem.get('y',0))
                w, h = float(elem.get('width',0)), float(elem.get('height',0))
                xs, ys = [x, x+w], [y, y+h]
            except: pass
        elif tag in ['polygon','polyline'] and 'points' in elem.attrib:
            nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', elem.get('points'))]
            xs, ys = nums[0::2], nums[1::2]
        elif tag == 'path':
            nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', elem.get('d',''))]
            xs, ys = nums[0::2], nums[1::2]
        
        if xs and ys:
            found = True
            min_x, max_x = min(min_x, min(xs)), max(max_x, max(xs))
            min_y, max_y = min(min_y, min(ys)), max(max_y, max(ys))
            
    if not found: return 0,0,100,100
    return min_x-10, min_y-10, (max_x-min_x)+20, (max_y-min_y)+20

def apply_offset(element, dx, dy):
    tag = element.tag.split('}')[-1]
    if tag == 'line':
        for a in ['x1','x2']: element.set(a, str(round(float(element.get(a,0))+dx, 2)))
        for a in ['y1','y2']: element.set(a, str(round(float(element.get(a,0))+dy, 2)))
    elif tag in ['rect','text']:
        element.set('x', str(round(float(element.get('x',0))+dx, 2)))
        element.set('y', str(round(float(element.get('y',0))+dy, 2)))
    elif tag == 'polygon' and 'points' in element.attrib:
        pts = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', element.get('points'))]
        new_pts = [f"{round(pts[i]+dx,2)},{round(pts[i+1]+dy,2)}" for i in range(0,len(pts),2)]
        element.set('points', " ".join(new_pts))
    elif tag == 'path':
        nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', element.get('d',''))]
        new_d = " ".join([f"{round(nums[i]+dx,2)},{round(nums[i+1]+dy,2)}" for i in range(0,len(nums),2)])
        element.set('d', f"M {new_d}" if new_d else "")
    if 'transform' in element.attrib: del element.attrib['transform']

def process_group_recursive(element, acc_x, acc_y, collector):
    curr_x, curr_y = acc_x, acc_y
    if 'transform' in element.attrib:
        match = re.search(r'translate\(\s*([-+]?\d*\.?\d+)\s*(?:[,\s]\s*([-+]?\d*\.?\d+))?\s*\)', element.get('transform'))
        if match:
            curr_x += float(match.group(1))
            curr_y += float(match.group(2)) if match.group(2) else 0.0
    
    if element.tag.endswith('g'):
        for child in list(element): process_group_recursive(child, curr_x, curr_y, collector)
    else:
        import copy
        new_elem = copy.deepcopy(element)
        apply_offset(new_elem, curr_x, curr_y)
        collector.append(new_elem)

def simplify_railroad_svg(svg_string):
    if not svg_string: return ""
    svg_string = svg_string.replace('&nbsp;', ' ')
    ET.register_namespace('', 'http://www.w3.org/2000/svg')
    try: root = ET.fromstring(svg_string)
    except: return svg_string
    
    new_root = ET.Element('svg')
    for k,v in root.attrib.items(): new_root.set(k,v)
    for child in root:
        if 'defs' in child.tag: new_root.append(child)
    
    flat_elements = []
    for child in root:
        if 'defs' not in child.tag: process_group_recursive(child, 0.0, 0.0, flat_elements)
    
    for elem in flat_elements: new_root.append(elem)
    x,y,w,h = calculate_bounds(new_root)
    new_root.set('viewBox', f"{x} {y} {w} {h}")
    new_root.set('width', f"{w}px")
    new_root.set('height', f"{h}px")
    return ET.tostring(new_root, encoding='unicode')

# ==========================================
# PART 3: CONNECT LOGIC (MAINLINE ONLY)
# ==========================================
def get_arrow_locations(root):
    arrows = []
    for elem in root.iter():
        if ('arrow' in elem.get('class','') or elem.tag.endswith('polygon')):
            pts = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', elem.get('points',''))]
            if len(pts)>=2: arrows.append((pts[0], pts[1]))
    return arrows

def add_red_lines(svg_string):
    flat_svg = simplify_railroad_svg(svg_string)
    flat_svg_clean = re.sub(r' xmlns="[^"]+"', '', flat_svg, count=1)
    try: root = ET.fromstring(flat_svg_clean)
    except: return flat_svg

    arrow_locs = get_arrow_locations(root)

    lines = []
    for elem in root.iter():
        if elem.tag.endswith('line'):
            try:
                y = float(elem.get('y1'))
                if abs(y - float(elem.get('y2'))) < 2.0:
                    lines.append({
                        'y': y, 
                        'min_x': min(float(elem.get('x1')), float(elem.get('x2'))), 
                        'max_x': max(float(elem.get('x1')), float(elem.get('x2')))
                    })
            except: continue
    
    if not lines: return flat_svg
    lines.sort(key=lambda k: k['y'])

    rows = []
    if lines:
        cluster = [lines[0]]
        for i in range(1, len(lines)):
            if abs(lines[i]['y'] - cluster[-1]['y']) < 8.0:
                cluster.append(lines[i])
            else:
                rows.append({
                    'y': sum(l['y'] for l in cluster)/len(cluster),
                    'end_x': max(l['max_x'] for l in cluster),
                    'start_x': min(l['min_x'] for l in cluster)
                })
                cluster = [lines[i]]
        rows.append({
            'y': sum(l['y'] for l in cluster)/len(cluster),
            'end_x': max(l['max_x'] for l in cluster),
            'start_x': min(l['min_x'] for l in cluster)
        })

    # FILTER: Mainlines Only (<60px start)
    mainline_rows = [r for r in rows if r['start_x'] < 60.0]

    # CONNECT
    for i in range(len(mainline_rows) - 1):
        curr = mainline_rows[i]
        next_row = mainline_rows[i+1]
        
        has_arrow = False
        for (ax, ay) in arrow_locs:
            if abs(curr['end_x'] - ax) < 20.0 and abs(curr['y'] - ay) < 25.0:
                has_arrow = True
                break
        
        if has_arrow:
            conn = ET.Element('line')
            conn.set('x1', str(curr['end_x']))
            conn.set('y1', str(curr['y']))
            conn.set('x2', str(next_row['start_x']))
            conn.set('y2', str(next_row['y']))
            conn.set('stroke', 'red')
            conn.set('stroke-width', '2')
            conn.set('class', 'row-connector')
            root.append(conn)

    return ET.tostring(root, encoding='unicode')

# ==========================================
# MAIN EXECUTION
# ==========================================
def main():
    print("=" * 70)
    print("SCRIPT 1: SMART LOCAL LOAD & PROCESS")
    print("=" * 70)

    # Make sure folders exist
    for folder in [RAW_DIR, REDLINES_DIR]:
        if not os.path.exists(folder):
            os.makedirs(folder)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Error: Input file '{INPUT_FILE}' not found.")
        return

    print(f"📂 Loading {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    driver = setup_driver()
    print("✅ Driver ready. Starting process...\n")

    processed_count = 0
    
    try:
        # Iterate rows starting from 2
        for row in range(2, ws.max_row + 1):
            cmd = ws.cell(row=row, column=1).value
            url = ws.cell(row=row, column=2).value
            
            if not cmd: break 
            print(f"[{row-1}] {cmd}...", end=" ")

            if not url:
                print("⚠️ Skipped (No URL)")
                continue

            raw_path = os.path.join(RAW_DIR, f"{cmd}.svg")
            conn_path = os.path.join(REDLINES_DIR, f"{cmd}.svg")
            raw_svg = None
            connected_svg = None

            # ---------------------------------------------------
            # STEP 1: LOAD OR FETCH RAW SVG
            # ---------------------------------------------------
            if os.path.exists(raw_path):
                with open(raw_path, 'r', encoding='utf-8') as f:
                    raw_svg = f.read()
                print("📂 Local Raw", end=" | ")
            else:
                raw_svg = fetch_svg_source_live(driver, url)
                if raw_svg:
                    with open(raw_path, 'w', encoding='utf-8') as f:
                        f.write(raw_svg)
                    print("🌐 Fetched", end=" | ")

            if not raw_svg:
                print("❌ Failed (No SVG)")
                continue

            # ---------------------------------------------------
            # STEP 2: LOAD OR GENERATE REDLINES SVG
            # ---------------------------------------------------
            if os.path.exists(conn_path):
                with open(conn_path, 'r', encoding='utf-8') as f:
                    connected_svg = f.read()
                print("📂 Local Redlines", end=" | ")
            else:
                connected_svg = add_red_lines(raw_svg)
                with open(conn_path, 'w', encoding='utf-8') as f:
                    f.write(connected_svg)
                print("⚙️ Processed", end=" | ")

            # ---------------------------------------------------
            # STEP 3: UPDATE EXCEL (CODE & IMAGES)
            # ---------------------------------------------------
            # A. Update Raw Code (Col C)
            ws.cell(row=row, column=3, value=raw_svg[:32000]).alignment = Alignment(wrap_text=True, vertical='top')

            # B. Screenshot Raw (Col D)
            img_path_raw = f"temp_raw_{row}.png"
            if take_screenshot(driver, raw_svg, img_path_raw):
                img = ExcelImage(img_path_raw)
                if img.height > 150:
                    ratio = 150/img.height
                    img.height = 150
                    img.width = int(img.width * ratio)
                ws.add_image(img, f"D{row}")

            # C. Update Connected Code (Col E)
            ws.cell(row=row, column=5, value=connected_svg[:32000]).alignment = Alignment(wrap_text=True, vertical='top')

            # D. Screenshot Connected (Col F)
            img_path_conn = f"temp_conn_{row}.png"
            if take_screenshot(driver, connected_svg, img_path_conn):
                img = ExcelImage(img_path_conn)
                if img.height > 150:
                    ratio = 150/img.height
                    img.height = 150
                    img.width = int(img.width * ratio)
                ws.add_image(img, f"F{row}")

            ws.row_dimensions[row].height = 150
            print("✅ Excel Updated")
            processed_count += 1

    finally:
        if driver: driver.quit()
        print(f"\n💾 Saving {INPUT_FILE} (This may take a moment)...")
        wb.save(INPUT_FILE)
        
        # Cleanup temp images
        if os.path.exists("temp_render.html"): os.remove("temp_render.html")
        for f in os.listdir():
            if f.startswith("temp_") and f.endswith(".png"):
                try: os.remove(f)
                except: pass

        print("-" * 70)
        print(f"🎉 Finished. Processed {processed_count} rows.")
        print(f"💾 Updated File: {INPUT_FILE}")

if __name__ == "__main__":
    main()