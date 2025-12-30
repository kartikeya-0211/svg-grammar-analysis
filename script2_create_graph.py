"""
SCRIPT 2: MASTER GRAPH GENERATOR (PRECISION MODE)
-------------------------------------------------
1. PRECISION BOXING: Padding reduced to 2px. No more "vacuuming" neighbors.
2. GARBAGE FILTER: Removes empty nodes like '()' or '((('.
3. RESCUE: Active.
4. ADAPTIVE SNAP: Active.
"""

import os
import re
import math
import copy
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# --- CONFIG ---
INPUT_FILE = 'railroad_diagrams.xlsx'
LINKS_FILE = 'links_cics.txt'
DRIVER_FILENAME = "msedgedriver.exe"

# --- CLASSES ---
class Node:
    def __init__(self, id, text, x, y, width=0, is_rect=False):
        self.id = id
        self.text = text
        self.x = float(x)
        self.y = float(y)
        self.width = float(width)
        self.left = self.x
        self.right = self.x + self.width if width > 0 else self.x + 30
        self.is_rect = is_rect 
    def __repr__(self): return f"{self.id}({self.text})"

class Edge:
    def __init__(self, start, end, type_):
        self.start = start
        self.end = end
        self.type = type_
    def __eq__(self, other):
        return self.start == other.start and self.end == other.end and self.type == other.type
    def __hash__(self): return hash((self.start, self.end, self.type))
    def __lt__(self, other): return (self.start, self.end) < (other.start, other.end)
    def __repr__(self): return f"{self.start}->{self.end}"

# --- SIMPLIFICATION ---
def simplify_svg_memory(svg_string):
    try:
        soup = BeautifulSoup(svg_string, "lxml-xml")
        clean_xml = str(soup).replace('&nbsp;', ' ')
        clean_xml = re.sub(r'(xmlns:?\w*="[^"]+")', '', clean_xml)
        clean_xml = re.sub(r'\w+:', '', clean_xml)
        root = ET.fromstring(clean_xml)
    except: return svg_string

    defs, main_group = None, None
    for child in root:
        if child.tag == 'defs': defs = child
        elif child.tag == 'g' and 'transform' in child.attrib: main_group = child
    
    new_root = ET.Element('svg')
    for k, v in root.attrib.items(): new_root.set(k, v)
    if defs: new_root.append(defs)

    flat_elems = []
    if main_group: process_group_recursive(main_group, 0.0, 0.0, flat_elems)
    for elem in flat_elems: new_root.append(elem)

    return ET.tostring(new_root, encoding='unicode')

def process_group_recursive(element, acc_x, acc_y, collector):
    curr_x, curr_y = acc_x, acc_y
    if 'transform' in element.attrib:
        m = re.search(r'translate\(\s*([-+]?\d*\.?\d+)\s*(?:[,\s]\s*([-+]?\d*\.?\d+))?\s*\)', element.get('transform'))
        if m:
            curr_x += float(m.group(1))
            curr_y += float(m.group(2)) if m.group(2) else 0.0
    
    if element.tag == 'g':
        for child in list(element): process_group_recursive(child, curr_x, curr_y, collector)
    else:
        new_elem = copy.deepcopy(element)
        apply_offset(new_elem, curr_x, curr_y)
        collector.append(new_elem)

def apply_offset(element, dx, dy):
    if element.tag == 'path':
        d = element.get('d', '')
        def shift(m):
            nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', m.group(0))]
            shifted = []
            for i in range(0, len(nums), 2):
                if i+1 < len(nums): shifted.append(f"{round(nums[i]+dx,2)},{round(nums[i+1]+dy,2)}")
            return " ".join(shifted)
        element.set('d', re.sub(r'[-+]?\d*\.?\d+[\s,]+[-+]?\d*\.?\d+', shift, d))
    elif element.tag in ['rect', 'text']:
        element.set('x', str(round(float(element.get('x', '0')) + dx, 2)))
        element.set('y', str(round(float(element.get('y', '0')) + dy, 2)))
    elif element.tag == 'line':
        for a in ['x1', 'y1', 'x2', 'y2']:
            val = float(element.get(a, '0'))
            element.set(a, str(round(val + (dx if 'x' in a else dy), 2)))
    if 'transform' in element.attrib: del element.attrib['transform']

# --- CORE PARSING ---
def calculate_distance(x1, y1, x2, y2):
    return math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

def get_nodes_from_svg(root):
    # 1. FIND RECTS (STRICT FILTER)
    rects = []
    for r in root.findall('.//rect'):
        try: 
            obj = {'x': float(r.get('x')), 'y': float(r.get('y')), 'w': float(r.get('width')), 'h': float(r.get('height', 20))}
            rects.append(obj)
        except: pass
    
    # Filter wrappers (>45px) and huge backgrounds
    valid_rects = [r for r in rects if r['w'] < 1000 and r['h'] < 45]
    valid_rects.sort(key=lambda r: (r['y'], r['x']))

    nodes = []
    consumed_texts = []

    # 2. EXTRACT TEXT
    all_text_elems = []
    for elem in root.iter():
        if 'text' in elem.tag:
            txt = (elem.text or "").strip()
            if not txt: continue
            try: 
                x, y = float(elem.get('x', 0)), float(elem.get('y', 0))
                all_text_elems.append({'txt': txt, 'x': x, 'y': y, 'obj': elem})
            except: continue

    # 3. MATCH TEXT TO RECTS (LASER PRECISION)
    for r in valid_rects:
        text_content = []
        for t in all_text_elems:
            # 🟢 FIX: Padding reduced to 2px to prevent grabbing neighbors
            if (r['x'] - 2 <= t['x'] <= r['x'] + r['w'] + 5) and \
               (r['y'] - 5 <= t['y'] <= r['y'] + r['h'] + 10):
                text_content.append(t)
                consumed_texts.append(t['obj'])
        
        if text_content:
            text_content.sort(key=lambda k: k['x'])
            full_text = " ".join([tc['txt'] for tc in text_content])
            # Clean formatting
            full_text = full_text.replace(" (", "(").replace("( ", "(").replace(" )", ")")
            nodes.append(Node("", full_text, r['x'], r['y'], r['w'], is_rect=True))

    # 4. PROCESS REMAINING TEXT
    bare_texts = [t for t in all_text_elems if t['obj'] not in consumed_texts]
    bare_texts.sort(key=lambda t: t['x'])
    
    if bare_texts:
        buffer, bx, by = [bare_texts[0]['txt']], bare_texts[0]['x'], bare_texts[0]['y']
        last_x_end = bare_texts[0]['x'] + (len(bare_texts[0]['txt']) * 8)
        
        for i in range(1, len(bare_texts)):
            curr = bare_texts[i]
            # 🟢 FIX: Strict Y-check (2px) and X-gap (3px)
            if abs(curr['y'] - by) < 2 and (curr['x'] - last_x_end) < 3:
                buffer.append(curr['txt'])
                last_x_end = curr['x'] + (len(curr['txt']) * 8)
            else:
                nodes.append(Node("", "".join(buffer), bx, by, last_x_end - bx, is_rect=False))
                buffer, bx, by = [curr['txt']], curr['x'], curr['y']
                last_x_end = curr['x'] + (len(curr['txt']) * 8)
        
        if buffer:
            nodes.append(Node("", "".join(buffer), bx, by, last_x_end - bx, is_rect=False))

    # 🟢 FIX: GARBAGE FILTER (Remove '()', '((', empty)
    clean_nodes = []
    idx = 1
    # Sort top-down, left-right for clean numbering
    nodes.sort(key=lambda n: (n.y, n.x))
    
    for n in nodes:
        # Reject nodes that are just punctuation
        if not re.search(r'[a-zA-Z0-9]', n.text):
            continue
        n.id = f"n{idx}"
        clean_nodes.append(n)
        idx += 1
        
    return clean_nodes

def get_closest_node(x, y, nodes, is_source, mainline_y, diagram_width=None):
    if not nodes: return None
    closest, min_dist = None, float('inf')

    for node in nodes:
        target_x = node.right if is_source else node.left
        target_y = node.y + (10 if node.is_rect else 0)
        
        dist = calculate_distance(x, y, target_x, target_y)
        if dist < min_dist: min_dist, closest = dist, node

    if diagram_width and diagram_width > 1500: THRESHOLD = 250
    elif diagram_width and diagram_width > 1000: THRESHOLD = 200
    else: THRESHOLD = 150
        
    return closest if min_dist < THRESHOLD else None

def classify_edge_type(start, end, points, mainline_y):
    start_x, _ = start
    end_x, _ = end
    y_vals = [pt[1] for pt in points]
    if end_x < (start_x - 20): return "Loopback"
    if max(y_vals) > (mainline_y + 15): return "Alternative"
    if min(y_vals) < (mainline_y - 15): return "Default"
    return "Mainline"

def parse_graph_data(svg_string):
    if not svg_string: return [], []
    try:
        soup = BeautifulSoup(svg_string, "lxml-xml")
        clean_xml = str(soup).replace('&nbsp;', ' ')
        clean_xml = re.sub(r'(xmlns:?\w*="[^"]+")', '', clean_xml)
        clean_xml = re.sub(r'\w+:', '', clean_xml)
        root = ET.fromstring(clean_xml)
    except: return [], []

    nodes = get_nodes_from_svg(root)
    if not nodes: return [], []
    
    mainline_y = nodes[0].y
    diagram_width = max(n.right for n in nodes) - min(n.left for n in nodes)

    raw_edges = []
    paths = []
    for elem in root.iter():
        if 'path' in elem.tag or 'line' in elem.tag: paths.append(elem)

    for p in paths:
        points = []
        tag = p.tag.split('}')[-1]
        if tag == 'path':
            d = p.get('d', '')
            nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', d)]
            for k in range(0, len(nums), 2):
                if k+1 < len(nums): points.append((nums[k], nums[k+1]))
        elif tag == 'line':
            try: points = [(float(p.get('x1')), float(p.get('y1'))), (float(p.get('x2')), float(p.get('y2')))]
            except: continue
        if len(points) < 2: continue

        start, end = points[0], points[-1]
        edge_type = classify_edge_type(start, end, points, mainline_y)
        src = get_closest_node(start[0], start[1], nodes, True, mainline_y, diagram_width)
        dst = get_closest_node(end[0], end[1], nodes, False, mainline_y, diagram_width)
        
        src_id, dst_id = (src.id if src else "START"), (dst.id if dst else "END")
        if src_id != dst_id or edge_type == "Loopback":
            raw_edges.append(Edge(src_id, dst_id, edge_type))

    return nodes, sorted(list(set(raw_edges)))

# --- MAIN ---
def setup_driver():
    if not os.path.exists(DRIVER_FILENAME): return None
    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument("--log-level=3") 
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    service = Service(executable_path=DRIVER_FILENAME)
    service.creation_flags = 0x08000000
    return webdriver.Edge(service=service, options=options)

def fetch_svg_direct(driver, url):
    try:
        driver.get(url)
        svg = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'svg.syntaxdiagram')))
        return svg.get_attribute('outerHTML')
    except: return None

def load_link_map(filename):
    if not os.path.exists(filename): return {}
    lm = {}
    with open(filename, 'r') as f:
        for l in f:
            url = l.strip()
            if not url: continue
            cmd = url.split('/')[-1].replace('.html', '').replace('dfhp4_', '').replace('dfhp4-', '')
            cmd_name = re.sub(r'([a-z])([A-Z])', r'\1 \2', cmd).upper()
            lm[cmd_name] = url
    return lm

def main():
    print("="*60)
    print("      SCRIPT 2: MASTER GRAPH GENERATOR (PRECISION)")
    print("="*60)

    if not os.path.exists(INPUT_FILE) or not os.path.exists(LINKS_FILE):
        print(f"❌ Missing {INPUT_FILE} or {LINKS_FILE}")
        return

    print("📖 Loading Links...")
    link_map = load_link_map(LINKS_FILE)
    
    try:
        wb = load_workbook(INPUT_FILE)
        ws = wb.active
        cell = ws.cell(row=1, column=7)
        cell.value = "Textual Representation simplified svg"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions['G'].width = 45
    except Exception as e:
        print(f"❌ Error: {e}")
        return

    driver = None
    row = 2
    count, rescued = 0, 0
    
    try:
        while True:
            cmd_name = ws.cell(row=row, column=1).value
            svg_code = ws.cell(row=row, column=5).value
            
            if not cmd_name: break
            
            is_broken = False
            if not svg_code or len(svg_code) > 32000 or not svg_code.strip().endswith('>'):
                is_broken = True

            final_svg = svg_code

            if is_broken:
                print(f"Row {row} ({cmd_name}): 🚨 BROKEN. Rescuing...", end="")
                if not driver: driver = setup_driver()
                url = link_map.get(cmd_name)
                if url:
                    fetched = fetch_svg_direct(driver, url)
                    if fetched:
                        final_svg = simplify_svg_memory(fetched)
                        print(" ✅ Rescued.")
                        rescued += 1
                    else:
                        print(" ❌ Fetch Failed.")
                else:
                    print(" ❌ No Link.")
            else:
                print(f"Row {row} ({cmd_name}): OK", end="")

            nodes, edges = parse_graph_data(final_svg)

            if not edges and nodes:
                print(" ⚠️  No edges.")
            elif not is_broken:
                print(" ✅")

            if nodes:
                output_text = "Nodes:\n"
                for n in nodes: output_text += f"  {n}\n"
                def get_edges(t): return [str(e) for e in edges if e.type == t]
                for label, t in [("Mainline edges", "Mainline"), ("Default edges", "Default"), 
                                 ("Alternative edges", "Alternative"), ("Loopback edges", "Loopback")]:
                    group = get_edges(t)
                    output_text += f"\n{label}:\n"
                    if group: output_text += "  " + "\n  ".join(group) + "\n"

                cell = ws.cell(row=row, column=7)
                cell.value = output_text
                cell.number_format = '@'
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            row += 1
            count += 1

    except KeyboardInterrupt: print("\n🛑 Stopped.")
    finally:
        if driver: driver.quit()
        wb.save(INPUT_FILE)
        print(f"\n✅ Done. Processed {count} graphs.")
        print(f"🚑 Rescued {rescued} truncated graphs.")

if __name__ == "__main__":
    main()