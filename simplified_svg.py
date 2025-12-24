"""
Formatter + Simplifier + Image Generator + AUTO CLEANUP
------------------------------------------------------------------------
1. Reads 'Original SVG' from Column B.
2. Formats it -> Column C.
3. Generates Image of Original -> Column D.
4. Simplifies it (Removes polygons, merges paths) -> Column E.
5. Generates Image of Simplified -> Column F.
6. DELETES all temporary images at the end to keep folder clean.
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
import xml.etree.ElementTree as ET
import os
import re
import time

# --- FILES ---
INPUT_FILE = 'railroad_diagrams.xlsx'
DRIVER_FILENAME = "msedgedriver.exe"


# 1. SETUP BROWSER (For Images)
def setup_driver():
    # Find the driver in the current folder
    cwd = os.getcwd()
    driver_path = os.path.join(cwd, DRIVER_FILENAME)
    
    if not os.path.exists(driver_path):
        print(f"❌ ERROR: '{DRIVER_FILENAME}' missing. Please put it in this folder.")
        return None

    print("Starting Edge Driver (Hidden)...")
    
    # Configure Edge options for headless (invisible) running
    options = Options()
    options.use_chromium = True
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--force-device-scale-factor=1')
    options.add_argument('--window-size=2000,2000') # Big canvas
    
    # Launch the driver service
    service = Service(executable_path=driver_path)
    return webdriver.Edge(service=service, options=options)

def svg_to_image(driver, svg_code, output_path):
    """Render SVG in browser, take screenshot, save to file."""
    if not svg_code: return False
    
    try:
        # Wrap SVG in clean HTML with white background
        html = f'<html><body style="margin: 0; padding: 20px; background: white;">{svg_code}</body></html>'
        
        # Save temp HTML file for the browser to load
        temp_html = os.path.abspath("temp_canvas.html")
        with open(temp_html, "w", encoding="utf-8") as f:
            f.write(html)
            
        # Open the file in Edge
        driver.get(f"file:///{temp_html}")
        time.sleep(0.5) # Wait for render
        
        # Find the SVG element and take a specific screenshot
        try:
            svg_element = driver.find_element(By.TAG_NAME, "svg")
            svg_element.screenshot(output_path)
            return True
        except:
            return False
    except Exception as e:
        print(f"     [Img Error]: {e}")
        return False

# 2. SIMPLIFICATION LOGIC (Your Rules)
# ==========================================
def simplify_railroad_svg(svg_string):
    ET.register_namespace('', 'http://www.w3.org/2000/svg')
    try:
        root = ET.fromstring(svg_string)
    except ET.ParseError:
        return svg_string

    # Filter: Keep only <defs> and main <g> transform group
    defs, main_group = None, None
    for child in root:
        tag = child.tag.split('}')[-1]
        if tag == 'defs': defs = child
        elif tag == 'g' and 'transform' in child.attrib: main_group = child
    
    # Rebuild the root with only essential parts
    if main_group is not None:
        for child in list(root): root.remove(child)
        if defs: root.append(defs)
        root.append(main_group)

    # Rule 1: Remove Arrowheads (Polygons)
    for parent in root.iter():
        for child in list(parent):
            if 'polygon' in child.tag:
                parent.remove(child)

    # Rule 2: Merge Text Nodes
    merge_text_nodes(root)

    # Rule 3: Merge Consecutive Paths
    merge_consecutive_paths(root)

    # Rule 4: Round Coordinates
    round_all_coordinates(root)

    # Clean Output String
    xml_str = ET.tostring(root, encoding='unicode')
    xml_str = re.sub(r'<ns\d+:', '<', xml_str)
    xml_str = re.sub(r'</ns\d+:', '</', xml_str)
    xml_str = re.sub(r'xmlns:ns\d+="[^"]*"', '', xml_str)
    
    return prettify_xml(xml_str)

def merge_text_nodes(root):
    # Combines adjacent text elements (like "WORD" + "(")
    for parent in root.iter():
        i = 0
        while i < len(parent) - 1:
            curr, next_node = parent[i], parent[i+1]
            if 'g' in curr.tag and 'g' in next_node.tag:
                t1 = curr.find('.//{http://www.w3.org/2000/svg}text') or curr.find('text')
                t2 = next_node.find('.//{http://www.w3.org/2000/svg}text') or next_node.find('text')
                if t1 is not None and t2 is not None:
                    t1.text = (t1.text or "") + (t2.text or "")
                    parent.remove(next_node)
                    continue
            i += 1

def merge_consecutive_paths(root):
    # Combines adjacent paths into one long path string
    for parent in root.iter():
        i = 0
        while i < len(parent) - 1:
            curr, next_node = parent[i], parent[i+1]
            if 'path' in curr.tag and 'path' in next_node.tag:
                d1, d2 = curr.get('d', ''), next_node.get('d', '')
                if d1 and d2:
                    curr.set('d', f"{d1} {d2}")
                    parent.remove(next_node)
                    continue
            i += 1

def round_all_coordinates(root):
    # Rounds all numbers to 2 decimal places
    for elem in root.iter():
        for attr in ['x','y','width','height','d','points','transform','rx','ry']:
            if attr in elem.attrib:
                val = elem.get(attr)
                elem.set(attr, re.sub(r'[-+]?\d*\.?\d+', lambda m: str(round(float(m.group(0)), 2)), val))

def prettify_xml(xml_str):
    # Adds proper indentation for readability
    xml_str = re.sub(r'>\s+<', '><', xml_str)
    lines = []
    indent = 0
    for part in re.split(r'(<[^>]+>)', xml_str):
        if not part.strip(): continue
        if '</' in part: indent -= 1
        lines.append('  ' * indent + part)
        if '<' in part and '</' not in part and '/>' not in part and '<?' not in part: indent += 1
    return '\n'.join(lines)

# ==========================================
# 3. MAIN EXECUTION
# ==========================================
def main():
    print("=" * 60)
    print("      FINAL PROCESSOR (TEXT + IMAGES + CLEANUP)")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"❌ Error: {INPUT_FILE} not found. Run the scraper first!")
        return

    # 1. Start Browser
    driver = setup_driver()
    if not driver: return

    # 2. Open Excel
    print(f"📂 Opening {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # 3. Process Rows
    row = 2
    count = 0
    
    print("\nProcessing...")
    try:
        while True:
            cmd = ws.cell(row=row, column=1).value
            raw_svg = ws.cell(row=row, column=2).value
            
            if not raw_svg: break
            
            print(f"[{count+1}] {cmd}: ", end="")

            # --- COLUMN C: Format Original ---
            ws.cell(row=row, column=3, value=prettify_xml(raw_svg)).alignment = Alignment(wrap_text=True, vertical='top')
            print("Fmt", end="..")

            # --- COLUMN D: Image Original ---
            img_name_1 = f"temp_orig_{row}.png"
            if svg_to_image(driver, raw_svg, img_name_1):
                ws.add_image(ExcelImage(img_name_1), f"D{row}")
                print("Img1", end="..")

            # --- COLUMN E: Simplify ---
            simple_svg = simplify_railroad_svg(raw_svg)
            ws.cell(row=row, column=5, value=simple_svg).alignment = Alignment(wrap_text=True, vertical='top')
            print("Simp", end="..")

            # --- COLUMN F: Image Simplified ---
            img_name_2 = f"temp_simp_{row}.png"
            if svg_to_image(driver, simple_svg, img_name_2):
                ws.add_image(ExcelImage(img_name_2), f"F{row}")
                print("Img2", end="..")

            # Adjust row height for images
            ws.row_dimensions[row].height = 120 
            
            print(" ✅ Done")
            row += 1
            count += 1

    except KeyboardInterrupt:
        print("\n⚠️  Stopped by user.")
        
    finally:
        # 4. Save and Cleanup
        driver.quit()
        print(f"\n💾 Saving {INPUT_FILE}...")
        wb.save(INPUT_FILE)
        
        # --- CLEANUP SECTION ---
        print("🧹 Cleaning up temporary files...")
        cleaned_count = 0
        
        # Remove temp HTML
        if os.path.exists("temp_canvas.html"): os.remove("temp_canvas.html")
        
        # Remove all PNGs starting with 'temp_'
        for filename in os.listdir():
            if filename.startswith("temp_") and filename.endswith(".png"):
                try:
                    os.remove(filename)
                    cleaned_count += 1
                except:
                    pass
                    
        print(f"   Removed {cleaned_count} temporary images.")
        
        print("-" * 60)
        print(f"✅ Finished! Processed {count} diagrams.")
        print("=" * 60)

if __name__ == "__main__":
    main()