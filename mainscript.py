"""
MASTER ORCHESTRATOR - FULL VISUAL & LOGICAL PIPELINE
=========================================================
1. READ (Local SVGs or fetch if missing) -> Column C
2. SCREENSHOT RAW (Calls Script 1) -> Column D
3. REDLINES/SIMPLIFY (Calls Script 1) -> Column E
4. SCREENSHOT CONNECTED (Calls Script 1) -> Column F
5. GRAMMAR MATH (Calls Script 2) -> Column G
6. REGEX MATH (Calls Script 3) -> Column H

DEPENDENCIES:
- pip install openpyxl selenium pillow
"""

import os
import sys
import re
import time
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

# --- STRICT FUNCTION IMPORTS ---
try:
    import script0_web_scraper as script0
    import script1_simplified_svg as script1
    import script2_unoptimized_grammar as script2
    import script3_optimized_regex as script3
except ImportError as e:
    print(f"❌ CRITICAL IMPORT ERROR: {e}")
    print("Ensure script0, script1, script2, and script3 are named correctly and in this folder.")
    sys.exit()

# --- CONFIG ---
LINKS_FILE = 'links_cics.txt'
OUTPUT_EXCEL = 'railroad_diagrams_complete.xlsx'
RAW_DIR = 'rawSVGs'
BLACKLINES_DIR = 'blacklinesSVGs'  

def resize_img(path, target_height=140):
    """Physically resizes image height so it fits perfectly inside the Excel row."""
    try:
        img = PILImage.open(path)
        w, h = img.size
        if h > target_height:
            scale = target_height / h
            img = img.resize((int(w * scale), target_height), PILImage.Resampling.LANCZOS)
            img.save(path)
    except Exception as e:
        print(f"  ⚠️ Resize warning: {e}")

def cleanup(path):
    """Silently deletes temporary files."""
    if os.path.exists(path):
        try: os.remove(path)
        except: pass

def main():
    print("=" * 75)
    print("🚀 MASTER PIPELINE: VISUALS + GRAMMAR + REGEX (MEMORY OPTIMIZED)")
    print("=" * 75)

    # 1. Setup Folders
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(BLACKLINES_DIR, exist_ok=True)

    if not os.path.exists(LINKS_FILE):
        print(f"❌ Error: '{LINKS_FILE}' missing.")
        return

    # 2. Get Limits
    with open(LINKS_FILE, 'r') as f:
        all_urls = [l.strip() for l in f if l.strip()]

    user_limit = input(f"Found {len(all_urls)} links. Process how many? (Enter number or 'all'): ").strip().lower()
    if user_limit == 'all':
        urls = all_urls
    else:
        try: urls = all_urls[:int(user_limit)]
        except: urls = all_urls[:3]

    # 3. Setup Browser (Needed for Screenshots)
    print("\n🌐 Starting WebDriver for Screenshots...")
    driver = script1.setup_driver()
    if not driver:
        print("❌ Failed to initialize WebDriver.")
        return

    # 4. Setup Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Railroad Data"

    headers = [
        "Command",              # A
        "URL",                  # B
        "Raw SVG Code",         # C
        "Raw SVG Image",        # D
        "Connected SVG Code",   # E
        "Connected SVG Image",  # F
        "Unoptimized Grammar",  # G
        "Optimized Regex"       # H
    ]
    
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Set Column Widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    for col in ['C', 'E', 'G', 'H']: ws.column_dimensions[col].width = 50
    for col in ['D', 'F']: ws.column_dimensions[col].width = 90

    temp_files = []
    row = 2
    processed_count = 0

    print("\n⚡ Processing Pipeline...")

    # 5. Execute Pipeline
    try:
        for idx, url in enumerate(urls, 1):
            cmd = script0.extract_command_name_from_url(url)
            print(f"[{idx}/{len(urls)}] {cmd:<18} ...", end=" ")
            
            raw_path = os.path.join(RAW_DIR, f"{cmd}.svg")
            conn_path = os.path.join(BLACKLINES_DIR, f"{cmd}.svg")

            # --- STEP A: READ OR FETCH RAW SVG ---
            raw_svg = ""
            if os.path.exists(raw_path):
                with open(raw_path, "r", encoding="utf-8") as f:
                    raw_svg = f.read()
            else:
                raw_svg = script1.fetch_svg_source_live(driver, url)
                if raw_svg:
                    with open(raw_path, "w", encoding="utf-8") as f:
                        f.write(raw_svg)

            if not raw_svg:
                print("⚠️ No SVG -> Skipped")
                ws.cell(row=row, column=1, value=cmd)
                ws.cell(row=row, column=2, value=url)
                ws.cell(row=row, column=3, value="[NO SVG DETECTED]")
                row += 1
                continue

            # --- STEP B: SCREENSHOT RAW SVG ---
            img_raw_path = f"temp_raw_{row}.png"
            has_raw_img = script1.take_screenshot(driver, raw_svg, img_raw_path)
            if has_raw_img: temp_files.append(img_raw_path)

            # --- STEP C: REDLINES & SIMPLIFY ---
            connected_svg = ""
            if os.path.exists(conn_path):
                with open(conn_path, 'r', encoding='utf-8') as f:
                    connected_svg = f.read()
            else:
                connected_svg = script1.add_red_lines(raw_svg)
                with open(conn_path, 'w', encoding='utf-8') as f:
                    f.write(connected_svg)

            # --- STEP D: SCREENSHOT CONNECTED SVG ---
            img_conn_path = f"temp_conn_{row}.png"
            has_conn_img = script1.take_screenshot(driver, connected_svg, img_conn_path)
            if has_conn_img: temp_files.append(img_conn_path)

            # --- STEP E: UNOPTIMIZED GRAMMAR ---
            grammar_text = ""
            try:
                # Fix namespaces and parse the connected SVG
                clean_svg = re.sub(r' xmlns="[^"]+"', '', connected_svg, count=1)
                root = ET.fromstring(clean_svg)
                for elem in root.iter():
                    if '}' in elem.tag: elem.tag = elem.tag.split('}')[1]
                
                blocks = script2.get_raw_blocks(root)
                paths, main_ys = script2.get_tracks_and_mainlines(root)
                grammar_text = script2.generate_grammar(blocks, paths, main_ys)
            except Exception as e:
                grammar_text = f"Error generating grammar: {e}"

            # --- STEP F: OPTIMIZED REGEX ---
            regex_text = ""
            try:
                if "Error" not in grammar_text and "n0 -> null" not in grammar_text:
                    regex_text = script3.optimize_to_regex(grammar_text)
                else:
                    regex_text = "$" if "n0 -> null" in grammar_text else "N/A"
            except Exception as e:
                regex_text = f"Error optimizing regex: {e}"

            # --- STEP G: WRITE TO EXCEL ROW ---
            ws.cell(row=row, column=1, value=cmd).alignment = Alignment(vertical='top')
            ws.cell(row=row, column=2, value=url).alignment = Alignment(vertical='top')
            
            ws.cell(row=row, column=3, value=raw_svg[:32000]).alignment = Alignment(wrap_text=True, vertical='top')
            if has_raw_img and os.path.exists(img_raw_path):
                resize_img(img_raw_path)
                ws.add_image(ExcelImage(img_raw_path), f"D{row}")

            ws.cell(row=row, column=5, value=connected_svg[:32000]).alignment = Alignment(wrap_text=True, vertical='top')
            if has_conn_img and os.path.exists(img_conn_path):
                resize_img(img_conn_path)
                ws.add_image(ExcelImage(img_conn_path), f"F{row}")

            ws.cell(row=row, column=7, value=grammar_text[:32000]).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row, column=8, value=regex_text[:32000]).alignment = Alignment(wrap_text=True, vertical='top')

            ws.row_dimensions[row].height = 150
            print("✅ Done")
            processed_count += 1
            row += 1

            # Save progress every 5 rows
            if idx % 5 == 0: 
                wb.save(OUTPUT_EXCEL)

    except KeyboardInterrupt:
        print("\n⚠️ Process interrupted by user. Saving current progress...")

    finally:
        print(f"\n💾 Saving final Excel file to '{OUTPUT_EXCEL}'...")
        wb.save(OUTPUT_EXCEL)
        if driver: driver.quit()
        
        # Clean up screenshot files
        for f in temp_files: cleanup(f)
        cleanup("temp_render.html")
        
        print(f"🎉 Master Pipeline Finished! Processed {processed_count} diagrams.")

if __name__ == "__main__":
    main()