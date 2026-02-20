import os
import re
import openpyxl
from openpyxl.styles import Alignment

def make_optional(r):
    if not r: return ""
    r = r.strip()
    
    if r.startswith('(') and r.endswith(')?'):
        depth = 0
        is_full = True
        for i, c in enumerate(r[:-2]):
            if c == '(': depth += 1
            elif c == ')': depth -= 1
            if depth == 0 and i > 0:
                is_full = False
                break
        if is_full: return r
        
    return f"({r})?"

def combine_seq(r1, r2):
    if r1 is None or r2 is None: return None
    if r1 == "": return r2
    if r2 == "": return r1
    return f"{r1} {r2}"

def make_star(r):
    if r is None or r == "": return ""
    r = r.strip()
    if len(r.split()) == 1 and not r.startswith('('): return f"{r}*"
    return f"({r})*"

def combine_or(r1, r2):
    if r1 is None: return r2
    if r2 is None: return r1
    if r1 == "" and r2 == "": return ""
    
    if r1 == "": return make_optional(r2)
    if r2 == "": return make_optional(r1)
    if r1 == r2: return r1
    
    p1, p2 = r1.split(), r2.split()
    common = 0
    while common < len(p1) and common < len(p2) and p1[common] == p2[common]:
        common += 1
        
    if common > 0:
        prefix = " ".join(p1[:common])
        rest = combine_or(" ".join(p1[common:]), " ".join(p2[common:]))
        if rest: return f"{prefix} {rest}".strip()
        return prefix
        
    return f"({r1}|{r2})"

def optimize_to_regex(grammar_text):
    if not grammar_text or grammar_text.strip() == "n0 -> null": return "$"
    
    graph = {}
    in_degree = {}
    
    for line in grammar_text.strip().split('\n'):
        if '->' not in line: continue
        parts = line.split('->')
        src = parts[0].strip()
        right = parts[1].strip().split()
        
        if len(right) >= 2:
            tgt = right[-1]
            label = " ".join(right[:-1])
        else:
            tgt, label = right[0], ""
            
        if label in ('null', '$'): label = ""
            
        if src not in graph: graph[src] = {}
        
        current_edge = graph[src].get(tgt, None)
        graph[src][tgt] = combine_or(current_edge, label)
            
        in_degree[tgt] = in_degree.get(tgt, 0) + 1
        if src not in in_degree: in_degree[src] = 0

    start_nodes = [n for n, deg in in_degree.items() if deg == 0 and n in graph]
    if start_nodes:
        start_nodes.sort(key=lambda x: int(x[1:]) if x[1:].isdigit() else 999)
        true_start = start_nodes[0]
    else:
        true_start = 'n0'
        
    accept_nodes = ['null', '$']
    for src, edges in graph.items():
        if not edges: accept_nodes.append(src)
        for tgt in list(edges.keys()):
            if tgt not in graph and tgt not in accept_nodes:
                accept_nodes.append(tgt)

    S, A = 'START', 'ACCEPT'
    graph[S] = {true_start: ""}
    
    for n in list(graph.keys()):
        for tgt in list(graph[n].keys()):
            if tgt in accept_nodes:
                label = graph[n].pop(tgt)
                graph[n][A] = combine_or(graph.get(n, {}).get(A, None), label)

    nodes_to_remove = [n for n in graph.keys() if n not in (S, A)]
    
    for node in nodes_to_remove:
        in_edges = [(s, graph[s][node]) for s in graph if node in graph[s]]
        out_edges = [(t, l) for t, l in graph.get(node, {}).items() if t != node]
        loop_regex = make_star(graph.get(node, {}).get(node, None))
        
        for src, in_label in in_edges:
            for tgt, out_label in out_edges:
                path_regex = combine_seq(in_label, combine_seq(loop_regex, out_label))
                current_edge = graph[src].get(tgt, None)
                graph[src][tgt] = combine_or(current_edge, path_regex)
                
        for src in graph:
            if node in graph[src]: del graph[src][node]
        if node in graph: del graph[node]
            
    final_regex = graph.get(S, {}).get(A, "")
    
    if final_regex:
        final_regex = final_regex.replace(" (", "(")
        if final_regex.endswith(' '): final_regex = final_regex.strip()
        
    if not final_regex.endswith('$'): 
        final_regex += '$'
        
    return final_regex

def process_script3_optimization(excel_filename):
    print("Starting Script 3: Optimizing Grammar to Regex")
    if not os.path.exists(excel_filename):
        print(f"Excel file '{excel_filename}' not found.")
        return
        
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    
    sheet.cell(row=1, column=8, value="Optimized Regex")

    for row in range(2, sheet.max_row + 1):
        raw_grammar = sheet.cell(row=row, column=7).value
        if not raw_grammar or raw_grammar == "SVG not found":
            continue
            
        optimized = optimize_to_regex(raw_grammar)
        target_cell = sheet.cell(row=row, column=8)
        target_cell.value = optimized
        target_cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(excel_filename)
    print("Optimization complete. Minimal Regex saved to Column H.")

if __name__ == "__main__":
    process_script3_optimization('railroad_diagrams.xlsx')