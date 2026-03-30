import os
import openpyxl

# ==========================================
# 1. AST BUILDERS & FACTORIZER (EPSILON EDITION)
# ==========================================
def create_node(label):
    if not label or label == 'null': return {'type': 'empty'}
    return {'type': 'text', 'val': label.strip()}

def make_star(node):
    if node is None or node['type'] == 'empty': return None
    if node['type'] == 'star': return node
    return {'type': 'star', 'inner': node}

def format_for_compare(node):
    if node is None or node['type'] == 'empty': return ""
    if node['type'] == 'epsilon': return "ε"
    if node['type'] == 'text': return node['val']
    if node['type'] == 'star': return f"*{format_for_compare(node['inner'])}*"
    if node['type'] == 'seq': return "".join(format_for_compare(p) for p in node['parts'])
    if node['type'] == 'or': 
        return "(" + "|".join(sorted(format_for_compare(p) for p in node['parts'])) + ")"

def make_seq(parts):
    flat = []
    for p in parts:
        if p is None or p.get('type') == 'empty': continue
        if p['type'] == 'seq': flat.extend(p['parts'])
        else: flat.append(p)
        
    # IBM Loop Collapse: A B (A B)* -> (A B)*
    reduced = []
    for p in flat:
        if p['type'] == 'star':
            if p['inner']['type'] == 'seq':
                star_parts = p['inner']['parts']
                star_len = len(star_parts)
                if len(reduced) >= star_len:
                    match = True
                    for i in range(star_len):
                        if format_for_compare(reduced[-star_len + i]) != format_for_compare(star_parts[i]):
                            match = False
                            break
                    if match:
                        for _ in range(star_len): reduced.pop() 
            elif len(reduced) > 0 and format_for_compare(reduced[-1]) == format_for_compare(p['inner']):
                reduced.pop()
                
        reduced.append(p)
        
    if not reduced: return {'type': 'empty'}
    if len(reduced) == 1: return reduced[0]
    return {'type': 'seq', 'parts': reduced}

def extract_seq(node):
    if node is None or node['type'] == 'empty': return []
    if node['type'] == 'seq': return node['parts']
    return [node]

def combine_or(node1, node2):
    if node1 is None: return node2
    if node2 is None: return node1

    is_empty1 = (node1['type'] == 'empty' or node1['type'] == 'epsilon')
    is_empty2 = (node2['type'] == 'empty' or node2['type'] == 'epsilon')
    if is_empty1 and is_empty2: return {'type': 'empty'}

    parts1 = node1['parts'] if node1['type'] == 'or' else [node1]
    parts2 = node2['parts'] if node2['type'] == 'or' else [node2]
    all_choices = parts1 + parts2
    
    seqs = [extract_seq(c) for c in all_choices]
    
    unique_seqs = []
    seen = set()
    for s in seqs:
        val = "".join(format_for_compare(n) for n in s)
        if val not in seen:
            seen.add(val)
            unique_seqs.append(s)
            
    if not unique_seqs: return {'type': 'empty'}
    
    # Check for empty paths to flag for Epsilon later
    has_empty = any(len(s) == 0 or (len(s)==1 and s[0].get('type')=='epsilon') for s in unique_seqs)
    non_empty = [s for s in unique_seqs if len(s) > 0 and s[0].get('type') != 'epsilon']
    
    if not non_empty: return {'type': 'epsilon'}
    if len(non_empty) == 1:
        core = make_seq(non_empty[0])
        # If optional, wrap in an OR with Epsilon
        return {'type': 'or', 'parts': [{'type': 'epsilon'}, core]} if has_empty else core
        
    # Prefix Factoring
    min_len = min(len(s) for s in non_empty)
    plen = 0
    for i in range(min_len):
        val = format_for_compare(non_empty[0][i])
        if all(format_for_compare(s[i]) == val for s in non_empty): plen += 1
        else: break
    prefix = non_empty[0][:plen]
    
    # Suffix Factoring
    after_prefix = [s[plen:] for s in non_empty]
    min_len_after = min(len(s) for s in after_prefix)
    slen = 0
    for i in range(1, min_len_after + 1):
        val = format_for_compare(after_prefix[0][-i])
        if all(format_for_compare(s[-i]) == val for s in after_prefix): slen += 1
        else: break
    suffix = after_prefix[0][-slen:] if slen > 0 else []
    
    cores = []
    for s in after_prefix:
        core = s[:-slen] if slen > 0 else s
        cores.append(make_seq(core))
        
    core_has_empty = any(c['type'] == 'empty' for c in cores)
    valid_cores = [c for c in cores if c['type'] != 'empty']
    
    if not valid_cores:
        or_node = {'type': 'empty'}
    elif len(valid_cores) == 1:
        or_node = valid_cores[0]
    else:
        or_node = {'type': 'or', 'parts': valid_cores}
        
    # Inject Epsilon explicitly if the branch is optional
    if core_has_empty or has_empty:
        if or_node['type'] == 'empty':
            or_node = {'type': 'epsilon'}
        elif or_node['type'] == 'or':
            or_node['parts'].insert(0, {'type': 'epsilon'})
        else:
            or_node = {'type': 'or', 'parts': [{'type': 'epsilon'}, or_node]}
        
    final_parts = prefix + ([or_node] if or_node['type'] != 'empty' else []) + suffix
    return make_seq(final_parts)

# ==========================================
# 2. AST FINAL FORMATTER (EPSILON STYLES)
# ==========================================
def format_ast(node):
    if node is None or node['type'] == 'empty': return ""
    if node['type'] == 'epsilon': return "ε"
    
    if node['type'] == 'text':
        return node['val']
        
    if node['type'] == 'star':
        inner = format_ast(node['inner']).strip()
        if not inner: return ""
        if node['inner']['type'] in ['seq', 'or']: return f"({inner})*"
        return f"{inner}*"
        
    if node['type'] == 'seq':
        texts = [format_ast(p).strip() for p in node['parts']]
        return " ".join([t for t in texts if t])
        
    if node['type'] == 'or':
        texts = []
        for p in node['parts']:
            val = format_ast(p).strip()
            if p['type'] == 'seq' and len(p['parts']) > 1: val = f"({val})"
            if val: texts.append(val)
            
        if not texts: return ""
        if len(texts) == 1: return texts[0]
        
        # Force Epsilon to the front, sort the rest by length
        if "ε" in texts:
            texts.remove("ε")
            texts.sort(key=len, reverse=True)
            texts.insert(0, "ε")
        else:
            texts.sort(key=len, reverse=True)
            
        return f"({' | '.join(texts)})"

# ==========================================
# 3. STATE ELIMINATION CORE
# ==========================================
def eliminate_states(text_block):
    graph = {}
    in_degree = {}
    all_nodes = set()

    for line in text_block.strip().split('\n'):
        if '->' not in line: continue
        parts = line.split('->')
        src, right = parts[0].strip(), parts[1].strip().split()
        
        if len(right) >= 2:
            label, tgt = " ".join(right[:-1]), right[-1]
        else:
            label, tgt = "", right[0]
            if tgt == 'null': tgt, label = 'ACCEPT', ''

        if src not in graph: graph[src] = {}
        all_nodes.update([src, tgt])
        in_degree[tgt] = in_degree.get(tgt, 0) + 1
        
        current_edge = graph[src].get(tgt, None)
        graph[src][tgt] = combine_or(current_edge, create_node(label))

    candidates = [n for n in all_nodes if in_degree.get(n, 0) == 0 and n != 'ACCEPT']
    def sort_key(n):
        return int(n[1:]) if n.startswith('n') and n[1:].isdigit() else 9999
    candidates.sort(key=sort_key)

    true_start = None
    for cand in candidates:
        visited, stack = set(), [cand]
        reaches_accept = False
        while stack:
            curr = stack.pop()
            if curr == 'ACCEPT':
                reaches_accept = True
                break
            if curr not in visited:
                visited.add(curr)
                for tgt in graph.get(curr, {}): stack.append(tgt)
        if reaches_accept:
            true_start = cand
            break
            
    if not true_start: true_start = candidates[0] if candidates else 'n0'
    
    if 'START' not in graph: graph['START'] = {}
    graph['START'][true_start] = {'type': 'empty'}

    def get_weight(node):
        in_count = sum(1 for s in graph if node in graph.get(s, {}))
        out_count = len(graph.get(node, {}))
        return in_count * out_count

    nodes_to_remove = [n for n in list(graph.keys()) + ['ACCEPT'] if n not in ('START', 'ACCEPT')]
    nodes_to_remove.sort(key=get_weight)

    for node in nodes_to_remove:
        in_edges = [(s, graph[s][node]) for s in graph if node in graph[s]]
        out_edges = [(t, l) for t, l in graph.get(node, {}).items() if t != node]
        
        loop_edge = graph.get(node, {}).get(node, None)
        loop_ast = make_star(loop_edge) if loop_edge else None

        for src, in_label in in_edges:
            for tgt, out_label in out_edges:
                bypass = make_seq([in_label, loop_ast, out_label])
                graph[src][tgt] = combine_or(graph[src].get(tgt, None), bypass)

        for src in list(graph.keys()):
            if node in graph[src]: del graph[src][node]
        if node in graph: del graph[node]

    final_ast = graph.get('START', {}).get('ACCEPT', None)
    return format_ast(final_ast)

# ==========================================
# 4. FULL EXCEL PIPELINE (FAST 2-PASS)
# ==========================================
def process_excel(excel_filename='railroad_diagrams.xlsx'):
    if not os.path.exists(excel_filename):
        print(f"Error: Could not find '{excel_filename}' in the current directory.")
        return

    print(f"Pass 1: Reading NFA grammar from {excel_filename} (Fast Mode)...")
    # Read-only mode skips loading heavy SVGs, preventing memory crashes
    wb_r = openpyxl.load_workbook(excel_filename, read_only=True, data_only=True)
    ws_r = wb_r.active
    
    row_data = []
    # Reads Column G (Index 7)
    for row in ws_r.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
        val = row[0]
        if val and '->' in str(val):
            row_data.append(str(val).strip())
        else:
            row_data.append(None)
    wb_r.close()
    
    print(f"Loaded {len(row_data)} valid rows. Processing math engine...")
    results = []
    for i, grammar in enumerate(row_data):
        if not grammar:
            results.append(None)
            continue
            
        try:
            raw_regex = eliminate_states(grammar)
            if not raw_regex: raw_regex = "(none)"
            
            # Cosmetic cleanup (removes messy spacing but retains the structural gaps)
            final_regex = raw_regex.replace(" (ε", "(ε") 
            final_regex = final_regex.replace(" )", ")").replace("( ", "(")
            final_regex = final_regex.replace("((", "( (").replace("))", ") )") 
            
            results.append(final_regex.strip())
        except Exception as e:
            print(f"Math Error on row {i+2}: {e}")
            results.append("(error)")
            
    print(f"Pass 2: Opening {excel_filename} to write results into Column H...")
    # Standard load to write data back
    wb_w = openpyxl.load_workbook(excel_filename)
    ws_w = wb_w.active
    
    # Write Header
    ws_w.cell(row=1, column=8).value = "Optimized Regex"
    
    # Write all compiled data to Column H (Index 8)
    for i, regex in enumerate(results):
        if regex is not None:
            ws_w.cell(row=i+2, column=8).value = regex
            
    print("Saving file... Please do not open it until finished.")
    wb_w.save(excel_filename)
    wb_w.close()
    
    print("Process Complete! All optimized regexes have been safely written to Column H.")

if __name__ == "__main__":
    process_excel()