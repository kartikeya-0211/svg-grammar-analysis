import os
import re
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# --- CONFIGURATION ---
LINKS_FILE = 'links_cics.txt'
RAW_SVGS_DIR = 'rawSVGs'
OUTPUT_EXCEL = 'railroad_diagrams.xlsx'
DRIVER_PATH = os.path.join(os.getcwd(), 'msedgedriver.exe')

def setup_driver():
    """Setup Edge driver."""
    edge_options = Options()
    edge_options.add_argument('--headless')
    edge_options.add_argument('--enable-unsafe-swiftshader')
    edge_options.add_argument("--log-level=3")
    edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    if not os.path.exists(DRIVER_PATH):
        raise Exception(f"❌ msedgedriver.exe not found at: {DRIVER_PATH}")
    
    service = Service(DRIVER_PATH)
    service.creation_flags = 0x08000000 
    
    return webdriver.Edge(service=service, options=edge_options)

def extract_command_name_from_url(url):
    """Extracts command name from URL."""
    filename = url.split('/')[-1].replace('.html', '')
    command = filename.replace('dfhp4_', '').replace('dfhp4-', '')
    words = re.sub(r'([a-z])([A-Z])', r'\1 \2', command).upper()
    return words.replace(" ", "_")

def extract_svg_from_web(driver, url):
    """Loads page and grabs fully-rendered SVG HTML."""
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "svg.syntaxdiagram"))
        )
        svg = driver.find_element(By.CSS_SELECTOR, "svg.syntaxdiagram")
        return svg.get_attribute("outerHTML")
    except Exception:
        return None

def prepare_excel():
    """Creates or Loads Excel and FORCES headers/formatting."""
    
    if os.path.exists(OUTPUT_EXCEL):
        print(f"📂 Loading existing Excel: {OUTPUT_EXCEL}")
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active
    else:
        print(f"📄 Creating new Excel: {OUTPUT_EXCEL}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Railroad Data"
    
    # --- FORCE HEADERS & WIDTHS (Every time) ---
    headers = [
        "Command",                  # A
        "URL",                      # B
        "Raw SVG Code",             # C
        "Raw SVG Image",            # D
        "Simplified SVG Code",      # E
        "Simplified SVG Image",     # F
        "Connected SVG Code",       # G
        "Connected SVG Image",      # H
        "Unoptimized Grammar",      # I
        "Optimized Regex"           # J
    ]
    
    # Apply Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply Column Widths
    # Standard
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    
    # Code Columns (Width 50)
    code_cols = ['C', 'E', 'G', 'I', 'J']
    for col_letter in code_cols:
        ws.column_dimensions[col_letter].width = 50
        
    # Image Columns (Width 90)
    image_cols = ['D', 'F', 'H']
    for col_letter in image_cols:
        ws.column_dimensions[col_letter].width = 90
    
    return wb, ws

def main():
    print("=" * 60)
    print("SCRIPT 0: DOWNLOAD SVGs + UPDATE EXCEL")
    print("=" * 60)

    # 1. Setup Folders
    if not os.path.exists(RAW_SVGS_DIR):
        os.makedirs(RAW_SVGS_DIR)
    
    if not os.path.exists(LINKS_FILE):
        print(f"❌ Error: '{LINKS_FILE}' not found.")
        return

    # 2. Prepare Excel (Forces Headers Now)
    wb, ws = prepare_excel()

    # 3. Read URLs
    with open(LINKS_FILE, 'r') as f:
        urls = [line.strip() for line in f if line.strip()]

    print(f"Processing {len(urls)} commands...")

    driver = None
    
    try:
        for i, url in enumerate(urls, 1):
            cmd_name = extract_command_name_from_url(url)
            file_path = os.path.join(RAW_SVGS_DIR, f"{cmd_name}.svg")
            
            # Row index (Headings are row 1, so data starts at i+1)
            current_row = i + 1
            
            # --- CHECK: Is this row already filled in Excel? ---
            # We check Column 1 (Command) and Column 3 (SVG Code)
            excel_has_cmd = ws.cell(row=current_row, column=1).value
            excel_has_svg = ws.cell(row=current_row, column=3).value
            
            if excel_has_cmd and excel_has_svg:
                # OPTIONAL: Uncomment to see skipped items
                # print(f"[{i}] {cmd_name} - Skipped (Done)")
                continue

            print(f"[{i}/{len(urls)}] {cmd_name}...", end=" ")

            svg_content = ""

            # --- STEP 1: GET SVG ---
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    svg_content = f.read()
                print("📂 Local File", end=" ")
            else:
                if driver is None: driver = setup_driver()
                svg_content = extract_svg_from_web(driver, url)
                if svg_content:
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(svg_content)
                    print("⬇️  Downloaded", end=" ")
                else:
                    print("❌ Not Found", end=" ")

            # --- STEP 2: UPDATE EXCEL ---
            ws.cell(row=current_row, column=1, value=cmd_name)
            ws.cell(row=current_row, column=2, value=url)
            
            if svg_content:
                cell = ws.cell(row=current_row, column=3, value=svg_content)
                # CRITICAL: This makes the text wrap and go DOWN
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                print("| ✅ Excel Updated")
            else:
                print("| ⚠️  No Data")

            # Save periodically
            if i % 10 == 0: wb.save(OUTPUT_EXCEL)

    finally:
        if driver: driver.quit()
        wb.save(OUTPUT_EXCEL)
        print("-" * 60)
        print(f"🎉 Process Complete.")
        print(f"📂 Files checked in '{RAW_SVGS_DIR}'")
        print(f"💾 Excel updated: '{OUTPUT_EXCEL}'")

if __name__ == "__main__":
    main()