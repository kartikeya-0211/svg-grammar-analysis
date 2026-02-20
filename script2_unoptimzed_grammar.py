import os
import re
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment

INPUT_EXCEL = 'railroad_diagrams.xlsx'
SVG_FOLDER = 'redlinesSVGs'

def get_raw_blocks(root):
    """Extracts rectangles and the text inside them."""
    blocks = []
    for rect in root.iter('rect'):
        x = float(rect.get('x', 0))
        y = float(rect.get('y', 0))
        w = float(rect.get('width', 0))
        h = float(rect.get('height', 0))
        
        text_str = ""
        for text in root.iter('text'):
            tx = float(text.get('x', 0))
            ty = float(text.get('y', 0))
            if x - 5 <= tx <= x + w + 5 and y - 10 <= ty <= y + h + 10:
                text_str += text.text.strip() if text.text else ""
        
        if text_str:
            blocks.append({'x': x, 'y': y, 'w': w, 'h': h, 'text': text_str})
    return blocks

def merge_split_words(blocks):
    """Stitches split parentheses or brackets back together."""
    blocks.sort(key=lambda b: (round(b['y'] / 10), b['x']))
    merged = []
    
    for b in blocks:
        if not merged:
            merged.append(b)
            continue
            
        prev = merged[-1]
        gap = b['x'] - (prev['x'] + prev['w'])
        
        if abs(prev['y'] - b['y']) < 10 and 0 <= gap <= 25:
            ptxt = prev['text'].strip()
            btxt = b['text'].strip()
            
            if btxt in ('(', ')') or ptxt.endswith('(') or btxt.startswith('(') or btxt.startswith(')'):
                prev['text'] += b['text']
                prev['w'] = (b['x'] + b['w']) - prev['x']
                continue
                
        merged.append(b)
            
    return merged

def get_tracks_and_mainlines(root):
    """Finds all line/path coordinates and horizontal mainlines."""
    paths = []
    main_ys = []
    
    for line in root.iter('line'):
        y1, y2 = float(line.get('y1', 0)), float(line.get('y2', 0))
        x1, x2 = float(line.get('x1', 0)), float(line.get('x2', 0))
        paths.append([(x1, y1), (x2, y2)])
        if abs(y1 - y2) < 2:
            main_ys.append(y1)
            
    for path in root.iter('path'):
        d = path.get('d', '')
        nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', d)]
        pts = [(nums[i], nums[i+1]) for i in range(0, len(nums)-1, 2)]
        if pts: paths.append(pts)
        
    if not main_ys: main_ys = [0.0]
    return paths, main_ys

def assign_branch_points(blocks, paths):
    """Traces the block's path back to where it split from the mainline."""
    for block in blocks:
        bx, by, bw, bh = block['x'], block['y'], block['w'], block['h']
        best_branch_x, best_branch_y = bx, by
        found = False
        
        for pts in paths:
            touches = any(bx - 30 <= px <= bx + 10 and by - 15 <= py <= by + bh + 15 for px, py in pts)
            if touches:
                start_x, start_y = pts[0]
                if not found or start_x < best_branch_x:
                    best_branch_x, best_branch_y = start_x, start_y
                    found = True
                    
        block['branch_x'] = best_branch_x
        block['branch_y'] = best_branch_y

def generate_grammar(blocks, paths, main_ys):
    if not blocks: return "n0 -> null"
    
    blocks = merge_split_words(blocks)
    assign_branch_points(blocks, paths)
    
    # 1. GROUP BY ROW
    blocks.sort(key=lambda b: b['branch_y'])
    rows, curr_row = [], []
    
    for b in blocks:
        if not curr_row:
            curr_row.append(b)
        else:
            avg_y = sum(c['branch_y'] for c in curr_row) / len(curr_row)
            if abs(b['branch_y'] - avg_y) < 25:
                curr_row.append(b)
            else:
                rows.append(curr_row)
                curr_row = [b]
    if curr_row: rows.append(curr_row)

    grammar = []
    n = 0
    
    # 2. PROCESS COLUMNS (Parallel Alternatives)
    for row in rows:
        row.sort(key=lambda b: b['branch_x'])
        columns, curr_col = [], []
        
        for b in row:
            if not curr_col:
                curr_col.append(b)
            else:
                avg_x = sum(c['branch_x'] for c in curr_col) / len(curr_col)
                if abs(b['branch_x'] - avg_x) < 25:
                    curr_col.append(b)
                else:
                    columns.append(curr_col)
                    curr_col = [b]
        if curr_col: columns.append(curr_col)
        
        avg_branch_y = sum(b['branch_y'] for b in row) / len(row)
        closest_main_y = min(main_ys, key=lambda my: abs(my - avg_branch_y))
        
        # 3. GENERATE LOGIC STRINGS
        for col in columns:
            curr_n, next_n = f"n{n}", f"n{n+1}"
            has_main = any(b['y'] - 2 <= closest_main_y <= b['y'] + b['h'] + 2 for b in col)
            
            if not has_main:
                grammar.append(f"{curr_n} -> {next_n}")
                
            for b in col:
                grammar.append(f"{curr_n} -> {b['text']} {next_n}")
                
            n += 1

    grammar.append(f"n{n} -> null")
    return "\n".join(grammar)

def process_railroad_diagrams(excel_filename, svg_folder):
    print("Starting Script 2: Generating Unoptimized Grammar")
    if not os.path.exists(excel_filename):
        print(f"Excel file '{excel_filename}' not found.")
        return
        
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    sheet.cell(row=1, column=7, value="Unoptimized Grammar")
    
    processed_count = 0

    for row in range(2, sheet.max_row + 1):
        raw_val = sheet.cell(row=row, column=1).value
        if not raw_val: break 
            
        cmd_name = str(raw_val).strip()
        absolute_path = os.path.abspath(os.path.join(svg_folder, f"{cmd_name}.svg"))
        
        target_cell = sheet.cell(row=row, column=7)
        target_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        if os.path.exists(absolute_path):
            try:
                tree = ET.parse(absolute_path)
                root = tree.getroot()
                
                for elem in root.iter():
                    if '}' in elem.tag: elem.tag = elem.tag.split('}')[1]
                
                blocks = get_raw_blocks(root)
                paths, main_ys = get_tracks_and_mainlines(root)
                target_cell.value = generate_grammar(blocks, paths, main_ys)
                
                print(f"Processed: {cmd_name}")
                processed_count += 1
            except Exception as e:
                target_cell.value = f"Error: {str(e)}"
        else:
            target_cell.value = "SVG not found"

    wb.save(excel_filename)
    print(f"Pipeline complete. Processed {processed_count} diagrams.")

if __name__ == "__main__":
    process_railroad_diagrams(INPUT_EXCEL, SVG_FOLDER)