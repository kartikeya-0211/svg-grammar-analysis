import os
import re
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment

INPUT_EXCEL = 'railroad_diagrams.xlsx'
SVG_FOLDER = 'blacklinesSVGs'
DEBUG_FILE = 'grammar_debug.txt' 

# ==========================================
# 1. SVG TRANSFORM PARSER (Absolute Math)
# ==========================================
def parse_transform(transform_str):
    """Extracts X,Y shifts from group tags like <g transform='translate(10,40)'>"""
    if not transform_str: return 0.0, 0.0
    match = re.search(r'translate\(\s*([-+]?\d*\.?\d+)\s*(?:,\s*([-+]?\d*\.?\d+))?\s*\)', transform_str)
    if match:
        dx = float(match.group(1))
        dy = float(match.group(2)) if match.group(2) else 0.0
        return dx, dy
    return 0.0, 0.0

def extract_absolute_elements(root):
    """Recursively walks the XML tree to calculate true global coordinates."""
    blocks = []
    paths = []

    def walk(node, current_dx, current_dy):
        dx, dy = parse_transform(node.get('transform'))
        abs_x = current_dx + dx
        abs_y = current_dy + dy

        if node.tag == 'rect':
            blocks.append({
                'type': 'rect',
                'x': float(node.get('x', 0)) + abs_x,
                'y': float(node.get('y', 0)) + abs_y,
                'w': float(node.get('width', 0)),
                'h': float(node.get('height', 0)),
                'text': ''
            })
            
        elif node.tag == 'text':
            if node.text and node.text.strip():
                blocks.append({
                    'type': 'text',
                    'x': float(node.get('x', 0)) + abs_x,
                    'y': float(node.get('y', 0)) + abs_y,
                    'val': node.text.strip()
                })
                
        elif node.tag == 'line':
            x1, y1 = float(node.get('x1', 0)) + abs_x, float(node.get('y1', 0)) + abs_y
            x2, y2 = float(node.get('x2', 0)) + abs_x, float(node.get('y2', 0)) + abs_y
            paths.append([(x1, y1), (x2, y2)])
            
        elif node.tag == 'path':
            d = node.get('d', '')
            nums = [float(n) for n in re.findall(r'[-+]?\d*\.?\d+', d)]
            pts = [((nums[i] + abs_x), (nums[i+1] + abs_y)) for i in range(0, len(nums)-1, 2)]
            if pts: paths.append(pts)

        for child in node:
            walk(child, abs_x, abs_y)

    walk(root, 0.0, 0.0)
    return blocks, paths

def map_text_to_blocks(raw_elements):
    """Assigns the floating text to their surrounding rectangles."""
    rects = [e for e in raw_elements if e['type'] == 'rect']
    texts = [e for e in raw_elements if e['type'] == 'text']
    
    final_blocks = []
    for r in rects:
        box_text = ""
        for t in texts:
            if r['x'] - 5 <= t['x'] <= r['x'] + r['w'] + 5 and r['y'] - 10 <= t['y'] <= r['y'] + r['h'] + 10:
                box_text += t['val']
        if box_text:
            final_blocks.append({'x': r['x'], 'y': r['y'], 'w': r['w'], 'h': r['h'], 'text': box_text})
            
    # Expanded Y-bucket to keep taller boxes in the correct visual row
    final_blocks.sort(key=lambda b: (round(b['y'] / 20), b['x']))
    
    merged = []
    for b in final_blocks:
        if not merged:
            merged.append(b)
            continue
        prev = merged[-1]
        gap = b['x'] - (prev['x'] + prev['w'])
        if abs(prev['y'] - b['y']) < 10 and 0 <= gap <= 25:
            ptxt, btxt = prev['text'].strip(), b['text'].strip()
            if btxt in ('(', ')') or ptxt.endswith('(') or btxt.startswith('(') or btxt.startswith(')'):
                prev['text'] += b['text']
                prev['w'] = (b['x'] + b['w']) - prev['x']
                continue
        merged.append(b)
    return merged

# ==========================================
# 2. THE EXACT COORDINATE GRAPH ALGORITHM
# ==========================================
def generate_grammar_from_coordinates(blocks, paths):
    """Maps the physical coordinates into an NFA Graph."""
    if not blocks: return "n0 -> null"
    
    nodes = []
    
    def get_node(x, y):
        for i, (nx, ny) in enumerate(nodes):
            # FIX: Expanded horizontal AND vertical snapping to safely grab detached arrows
            if abs(nx - x) <= 35 and abs(ny - y) <= 20: 
                return f"n{i}"
        nodes.append((x, y))
        return f"n{len(nodes)-1}"

    grammar_edges = set()
    
    for b in blocks:
        center_y = b['y'] + (b['h'] / 2)
        n_start = get_node(b['x'], center_y)
        n_end = get_node(b['x'] + b['w'], center_y)
        grammar_edges.add((n_start, b['text'], n_end))
        
    for pts in paths:
        if not pts: continue
        x1, y1 = pts[0]
        x2, y2 = pts[-1]
        n_start = get_node(x1, y1)
        n_end = get_node(x2, y2)
        if n_start != n_end:
            grammar_edges.add((n_start, "", n_end))
            
    # Connect dead ends to null
    outgoing = {s for s, l, e in grammar_edges}
    incoming = {e for s, l, e in grammar_edges}
    
    for n in range(len(nodes)):
        node_name = f"n{n}"
        if node_name not in outgoing and node_name in incoming:
            grammar_edges.add((node_name, "", "null"))
            
    grammar_strings = []
    for s, l, e in grammar_edges:
        if l: grammar_strings.append(f"{s} -> {l} {e}")
        else: grammar_strings.append(f"{s} -> {e}")
        
    return "\n".join(sorted(grammar_strings))

# ==========================================
# 3. PIPELINE EXECUTION
# ==========================================
def process_railroad_diagrams(excel_filename, svg_folder):
    print("Starting Script 2: Coordinate-Mapped NFA Generator")
    if not os.path.exists(excel_filename):
        print(f"Excel file '{excel_filename}' not found.")
        return
        
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
    sheet.cell(row=1, column=7, value="Unoptimized Grammar")
    
    processed_count = 0

    print(f"📝 Logging output to {DEBUG_FILE}...")
    with open(DEBUG_FILE, 'w', encoding='utf-8') as log_file:
        log_file.write("=== SCRIPT 2: ABSOLUTE COORDINATE GRAPH LOG ===\n\n")

        for row in range(2, sheet.max_row + 1):
            raw_val = sheet.cell(row=row, column=1).value
            if not raw_val: break 
                
            cmd_name = str(raw_val).strip()
            absolute_path = os.path.abspath(os.path.join(svg_folder, f"{cmd_name}.svg"))
            target_cell = sheet.cell(row=row, column=7)
            target_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if os.path.exists(absolute_path):
                try:
                    tree = ET.parse(absolute_path)
                    root = tree.getroot()
                    
                    for elem in root.iter():
                        if '}' in elem.tag: elem.tag = elem.tag.split('}')[1]
                    
                    raw_elements, paths = extract_absolute_elements(root)
                    blocks = map_text_to_blocks(raw_elements)
                    grammar_result = generate_grammar_from_coordinates(blocks, paths)
                    
                    target_cell.value = grammar_result
                    log_file.write(f"--- COMMAND: {cmd_name} ---\n{grammar_result}\n\n")
                    processed_count += 1
                except Exception as e:
                    target_cell.value = f"Error: {str(e)}"
            else:
                target_cell.value = "SVG not found"

    wb.save(excel_filename)
    print(f"Pipeline complete. Processed {processed_count} diagrams.")

if __name__ == "__main__":
    process_railroad_diagrams(INPUT_EXCEL, SVG_FOLDER)