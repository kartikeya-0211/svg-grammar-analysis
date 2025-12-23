import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook

def simplify_railroad_svg(svg_string):
    """
    Parses and simplifies an SVG railroad diagram for grammar analysis.
    
    Transformation Steps:
    1. Parsing: Loads SVG and registers namespaces.
    2. Filtering: Retains only <defs>, <style>, and the main <g> content group.
    3. Cleaning: Removes decorative <polygon> elements (arrowheads).
    4. Normalization: Converts curved paths (Q) into orthogonal rectilinear paths (L).
    5. Consolidation: Merges adjacent syntax text nodes into single strings.
    6. Formatting: Rounds coordinates and standardizes XML indentation.
    """
    
    # Register namespace to prevent 'ns0:' prefixes in output
    ET.register_namespace('', 'http://www.w3.org/2000/svg')
    
    try:
        root = ET.fromstring(svg_string)
    except ET.ParseError:
        return svg_string
    
    # --- Step 1: Filter Main Content ---
    defs_element = None
    main_group = None
    
    # Identify critical structural elements
    for child in root:
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'defs':
            defs_element = child
        elif tag == 'g' and 'transform' in child.attrib:
            main_group = child
    
    if main_group is None:
        return svg_string  # Return original if specific structure isn't found
    
    # Reconstruct root with only essential elements
    for child in list(root):
        root.remove(child)
    
    if defs_element is not None:
        root.append(defs_element)
    root.append(main_group)
    
    # --- Step 2: Remove Artifacts ---
    remove_polygons(root)
    
    # --- Step 3: Simplify Geometry ---
    convert_paths_to_orthogonal(root)
    
    # --- Step 4: Merge Text Tokens ---
    merge_text_nodes(root)
    
    # --- Step 5: Standardization ---
    round_all_coordinates(root)
    
    # Generate Output String
    xml_str = ET.tostring(root, encoding='unicode')
    
    # Clean up namespace artifacts
    xml_str = re.sub(r'<ns\d+:', '<', xml_str)
    xml_str = re.sub(r'</ns\d+:', '</', xml_str)
    xml_str = re.sub(r'xmlns:ns\d+="[^"]*"', '', xml_str)
    
    return prettify_xml(xml_str)


def remove_polygons(root):
    """Removes <polygon> tags, typically used for arrowheads."""
    for parent in root.iter():
        for child in list(parent):
            tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
            if tag == 'polygon':
                parent.remove(child)


def convert_paths_to_orthogonal(root):
    """
    Iterates through all <path> elements and converts Quadratic curves (Q)
    into orthogonal Line (L) segments (90-degree corners).
    """
    for path in root.iter():
        tag = path.tag.split('}')[-1] if '}' in path.tag else path.tag
        if tag == 'path':
            d_attr = path.get('d', '')
            if d_attr:
                new_d = make_orthogonal_path(d_attr)
                path.set('d', new_d)


def make_orthogonal_path(path_d):
    """
    Transforms path data: 'Q cx cy x y' -> 'L cx cy L x y'.
    This forces the path to travel through the control point as a corner.
    """
    tokens = re.findall(r'([MQL])\s*((?:[-+]?\d*\.?\d+\s*[, ]?\s*)+)', path_d)
    
    if not tokens:
        return path_d
    
    result_parts = []
    
    for cmd, coords_str in tokens:
        coords = [float(x) for x in re.findall(r'[-+]?\d*\.?\d+', coords_str)]
        
        if cmd == 'M': # Move
            if len(coords) >= 2:
                x, y = round(coords[-2], 2), round(coords[-1], 2)
                result_parts.append(f'M{x},{y}')
                
        elif cmd == 'L': # Line
            if len(coords) >= 2:
                x, y = round(coords[-2], 2), round(coords[-1], 2)
                result_parts.append(f'L{x},{y}')
                
        elif cmd == 'Q': # Quadratic Curve
            # Convert Curve to Corner: Line to control point -> Line to end point
            if len(coords) >= 4:
                cx, cy = round(coords[0], 2), round(coords[1], 2)
                x, y = round(coords[2], 2), round(coords[3], 2)
                result_parts.append(f'L{cx},{cy}')
                result_parts.append(f'L{x},{y}')
    
    return ' '.join(result_parts)


def merge_text_nodes(root):
    """
    Consolidates split syntax tokens.
    Merges <text class="syntaxkwd"> with immediately following 
    <text class="syntaxvar"> or <text class="syntaxdelim"> nodes.
    """
    for parent in root.iter():
        i = 0
        # Use a while loop to handle list modification during iteration
        while i < len(parent) - 1:
            current = parent[i]
            next_node = parent[i+1]
            
            def get_info(node):
                tag = node.tag.split('}')[-1] if '}' in node.tag else node.tag
                cls = node.get('class', '')
                return tag, cls
            
            curr_tag, curr_cls = get_info(current)
            next_tag, next_cls = get_info(next_node)
            
            # Check if both elements are groups containing text
            if curr_tag == 'g' and next_tag == 'g' and 'text' in curr_cls and 'text' in next_cls:
                
                # Locate the internal <text> tag within the grouping <g> tag
                curr_text_el = current.find('.//{http://www.w3.org/2000/svg}text')
                if curr_text_el is None: curr_text_el = current.find('text')
                
                next_text_el = next_node.find('.//{http://www.w3.org/2000/svg}text')
                if next_text_el is None: next_text_el = next_node.find('text')
                
                if curr_text_el is not None and next_text_el is not None:
                    # Merge content: Append next text to current text
                    curr_text_el.text = (curr_text_el.text or "") + (next_text_el.text or "")
                    
                    # Remove the redundant node
                    parent.remove(next_node)
                    
                    # Continue without incrementing 'i' to check for further merges
                    continue

            i += 1


def round_all_coordinates(root):
    """Rounds all numerical attributes to 2 decimal places for consistency."""
    for elem in root.iter():
        for attr in ['x', 'y', 'x1', 'y1', 'x2', 'y2', 'width', 'height', 'd', 'points', 'transform']:
            if attr in elem.attrib:
                val = elem.get(attr)
                new_val = re.sub(r'[-+]?\d*\.?\d+', lambda m: str(round(float(m.group(0)), 2)), val)
                elem.set(attr, new_val)


def prettify_xml(xml_str):
    """Formats the XML string with proper indentation."""
    xml_str = re.sub(r'>\s+<', '><', xml_str)
    lines = []
    indent = 0
    for part in re.split(r'(<[^>]+>)', xml_str):
        if not part.strip(): continue
        if part.startswith('</'):
            indent -= 1
            lines.append('  ' * indent + part)
        elif part.startswith('<') and not part.endswith('/>'):
            lines.append('  ' * indent + part)
            if not part.startswith('<?') and not part.startswith('<!'):
                indent += 1
        else:
            lines.append('  ' * indent + part)
    return '\n'.join(lines)
def process_excel_file(filename):
    """Main execution flow: Reads Excel, processes SVGs, and saves results."""
    print("=" * 60)
    print(f"📂 Processing File: {filename}")
    print("=" * 60)
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except Exception as e:
        print(f"❌ Critical Error: Could not open file. Details: {e}")
        return
    count = 0
    row_num = 1
    while True:
        # Read Original SVG from Column A
        cell = ws.cell(row=row_num, column=1)
        original_svg = cell.value
        # Stop at the first empty row
        if not original_svg:
            break
        # Process only if valid SVG content is detected
        if isinstance(original_svg, str) and '<svg' in original_svg:
            try:
                clean_svg = simplify_railroad_svg(original_svg)
                # Write Simplified SVG to Column B
                ws.cell(row=row_num, column=2, value=clean_svg)
                count += 1
            except Exception as e:
                ws.cell(row=row_num, column=2, value=f"Error: {e}")
        
        row_num += 1
    wb.save(filename)
    print(f"✅ Completed. {count} diagrams processed.")
    print(f"💾 Results saved to Column B in: {filename}")

if __name__ == "__main__":
    # Define input file
    EXCEL_FILE = 'railroad_diagrams.xlsx'
    process_excel_file(EXCEL_FILE)
    print("\nPress Enter to exit...")
    input()